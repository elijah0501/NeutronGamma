import openpyxl
import torch
from torch.autograd import Variable

from alexnet import AlexNet
from dataloader.data_loader import create_prediction_dataloader

weight_path = './model.pkl'

device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
print("Using" + str(device))

net = AlexNet().to(device)
print(net)

file = "data/real_data.xlsx"
dataloader = create_prediction_dataloader(file, 'real_data')

net.load_state_dict(torch.load(weight_path))
prediction = list()

for data, label in dataloader:
    X = torch.unsqueeze(data, dim=1)

    X = X.to(device)
    label = label.to(device)
    # 转换为Variable类型
    X = Variable(X)
    Y = Variable(label)

    # feedforward
    output = net(X)

    maxvalue, predict = torch.max(output, 1)
    print('the max value is ')
    print(maxvalue)
    predict = predict.to('cpu').data.numpy()
    print('the predict index is')
    print(predict)
    if predict == 0:
        prediction.append('Gamma')
    elif predict == 1:
        prediction.append("Neutron")

prediction_workbook = openpyxl.Workbook()
prediction_worksheet = prediction_workbook.active
prediction_worksheet.title = "prediction"

real_data_wb = openpyxl.load_workbook(file)
real_data_st = real_data_wb.get_sheet_by_name('real_data')


def get_sheet_data(sheet, prediction_label):
    sheet_data = list()
    i = 0
    for columns in sheet:
        a = list()
        counter = 0
        for cell in columns:
            a.append(int(cell.value))
            counter += 1
            if counter > 99:
                a.append(prediction_label[i])
        sheet_data.append(a)
        i += 1
    return sheet_data


prediction_list = get_sheet_data(real_data_st, prediction)


def write_xlsx(worksheet, sheet_data):
    i = 1
    for row in range(6000):
        j = 1
        for column in range(101):
            worksheet.cell(i, j).value = sheet_data[i - 1][j - 1]
            j += 1
        i += 1


write_xlsx(real_data_st, prediction_list)
real_data_wb.save(filename='./data/prediction.xlsx')
