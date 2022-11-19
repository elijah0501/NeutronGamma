import openpyxl
import random
import numpy as np
import scipy.io as scio


def create_xlsx(sheet_name):
    workbook_v = openpyxl.Workbook()
    sheet_v = workbook_v.active
    sheet_v.title = sheet_name
    return workbook_v, sheet_v


def write_xlsx(worksheet, sheet_data):
    i = 1
    for row in range(len(sheet_data)):
        j = 1
        for column in range(len(sheet_data[0])):
            worksheet.cell(i, j).value = sheet_data[i - 1][j - 1]
            j += 1
        i += 1


def open_xlsx(path, sheet_name):
    print(sheet_name)
    workbook_v = openpyxl.load_workbook(path)
    sheet_v = workbook_v.get_sheet_by_name(sheet_name)

    return workbook_v, sheet_v


def read_excel(file, sheet_name):
    _, st = open_xlsx(file, sheet_name)
    sheet_data = np.array(get_sheet_data_by_column(st))

    return sheet_data


def get_sheet_data_by_column(sheet):
    """
    Read sheet data from sheet object only
    :param sheet: the sheet object
    :param sheet_data: an empty list for containing sheet data
    :return: a list with all sheet data
    """
    sheet_data = list()
    for columns in sheet:
        st_column = list()
        for cell in columns:
            st_column.append(float(cell.value))
        sheet_data.append(st_column)

    return sheet_data


def get_sheet_data_by_rows(sheet, sheet_data):
    """
    Read sheet data from sheet object only
    :param sheet: the sheet object
    :param sheet_data: an empty list for containing sheet data
    :return: a list with all sheet data
    """
    for row in sheet.rows:
        rows = list()
        for cell in row:
            rows.append(float(cell.value))
        sheet_data.append(rows)

    return sheet_data


def label_data(sheet_data, label):
    labeled_data = list()
    for row in range(len(sheet_data)):
        st_column = list()
        for item in range(len(sheet_data[0])):
            st_column.append(sheet_data[row][item])
        st_column.append(label)
        labeled_data.append(st_column)
    return labeled_data


def label_data_without_norm(sheet, label):
    labeled_data = list()
    for columns in sheet:
        st_column = list()
        for cell in columns:
            st_column.append(float(cell.value))
        st_column.append(float(label))
        labeled_data.append(st_column)
    return labeled_data


def get_mixed_data(label_list):
    data_len = len(label_list)
    indices = list(range(data_len))
    random.shuffle(indices)
    mixed_data = []
    for index in indices:
        mixed_data.append(label_list[index])
    return mixed_data


def load_mat_data(path, flag, dict_name):
    data_dict = scio.loadmat(path)
    if flag:
        data_dict = data_dict[dict_name] * (-1)
    else:
        data_dict = data_dict[dict_name]

    return data_dict
